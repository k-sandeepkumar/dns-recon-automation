from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def format_sheet(ws):
    """
    Apply basic formatting to worksheet.
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center"
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column in ws.columns:
        max_length = 0

        column_letter = get_column_letter(
            column[0].column
        )

        for cell in column:
            try:
                value_length = len(str(cell.value))
                max_length = max(max_length, value_length)
            except Exception:
                pass

        ws.column_dimensions[
            column_letter
        ].width = min(max_length + 2, 60)


def add_simple_sheet(workbook, sheet_name, values):
    """
    Create a sheet containing a single-column list.
    """

    ws = workbook.create_sheet(sheet_name)

    ws.append(["Subdomain"])

    for value in values:
        ws.append([value])

    format_sheet(ws)

    return ws


def add_combined_sheet(workbook, values):
    """
    Create combined subdomain sheet.
    """

    ws = workbook.create_sheet("Combined")

    ws.append([
        "Subdomain",
        "Source"
    ])

    for subdomain, sources in values.items():
        ws.append([
            subdomain,
            ", ".join(sorted(sources))
        ])

    format_sheet(ws)

    return ws


def add_dns_sheet(workbook, dns_records):
    """
    Create DNS results sheet.
    """

    ws = workbook.create_sheet("DNSx")

    ws.append([
        "Host",
        "DNS Result"
    ])

    for host, result in dns_records:
        ws.append([
            host,
            result
        ])

    format_sheet(ws)

    return ws


def add_http_sheet(workbook, http_results):
    """
    Create HTTP results sheet.
    """

    ws = workbook.create_sheet("HTTPx")

    if not http_results:
        ws.append(["URL"])
        format_sheet(ws)
        return ws

    headers = [
        "URL",
        "Status",
        "Title",
        "Technology",
        "Web Server",
        "IP"
    ]

    ws.append(headers)

    for result in http_results:
        ws.append([
            result.get("url", ""),
            result.get("status", ""),
            result.get("title", ""),
            result.get("technology", ""),
            result.get("webserver", ""),
            result.get("ip", "")
        ])

    format_sheet(ws)

    return ws


def create_workbook(
    output_file,
    tool_results,
    combined_results,
    dns_results,
    http_results
):
    """
    Create the complete Excel workbook.
    """

    workbook = Workbook()

    # Remove default worksheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Individual tool sheets
    for tool_name, values in tool_results.items():
        add_simple_sheet(
            workbook,
            tool_name,
            values
        )

    # Combined
    add_combined_sheet(
        workbook,
        combined_results
    )

    # DNS
    add_dns_sheet(
        workbook,
        dns_results
    )

    # HTTP
    add_http_sheet(
        workbook,
        http_results
    )

    workbook.save(output_file)